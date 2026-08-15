#!/usr/bin/env python3
"""Importa a aba 'base' de uma planilha para o fluxo de Lançamentos do Nossas Finanças.

O importador reproduz a lógica de src/app/lancamentos/page.tsx:
- cria um registro em lancamentos;
- para CONTA, cria uma transacao paga;
- para CARTAO, cria/usa a fatura correspondente e cria uma transacao por parcela.

Por segurança, o comando padrão é somente validação (dry-run). Para gravar no Supabase,
use --executar.

Variáveis de ambiente obrigatórias para execução:
    SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY

Opcional:
    IMPORT_CRIADO_POR_NOME (padrão: "Importação")

Exemplo:
    python scripts/importar_lancamentos.py "carga_base_dados(1).xlsx"
    python scripts/importar_lancamentos.py "carga_base_dados(1).xlsx" --executar
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError:
    print("Erro: instale a dependência do importador com: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

EXPECTED_COLUMNS = [
    "data",
    "descricao",
    "valor",
    "tipo",
    "tipo_pessoa",
    "conta",
    "categoria",
    "meio_pagamento",
    "cartao",
    "parcelas",
]


class ImportErrorDetalhado(Exception):
    pass


@dataclass
class Categoria:
    id: str
    nome: str
    tipo: str
    parent_id: str | None


@dataclass
class RegistroEntrada:
    linha: int
    data: str
    descricao: str
    valor: Decimal
    tipo: str
    tipo_pessoa: str
    conta: str | None
    categoria: str
    meio_pagamento: str
    cartao: str | None
    parcelas: int


class SupabaseREST:
    def __init__(self, url: str, service_role_key: str):
        self.base_url = url.rstrip("/") + "/rest/v1"
        self.headers = {
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
        }

    def request(
        self,
        method: str,
        table: str,
        *,
        params: dict[str, str] | None = None,
        payload: Any = None,
        prefer: str | None = None,
    ) -> Any:
        url = f"{self.base_url}/{table}"
        if params:
            url += "?" + urlencode(params)

        headers = dict(self.headers)
        if prefer:
            headers["Prefer"] = prefer

        body = None if payload is None else __import__("json").dumps(payload).encode("utf-8")
        request = Request(url, data=body, headers=headers, method=method)

        try:
            with urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8")
                if not raw:
                    return None
                return __import__("json").loads(raw)
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ImportErrorDetalhado(
                f"Supabase respondeu HTTP {exc.code} em {method} {table}: {detail}"
            ) from exc
        except URLError as exc:
            raise ImportErrorDetalhado(f"Não foi possível acessar o Supabase: {exc.reason}") from exc

    def select(self, table: str, columns: str = "*", filters: dict[str, str] | None = None) -> list[dict[str, Any]]:
        params = {"select": columns}
        if filters:
            params.update(filters)
        result = self.request("GET", table, params=params)
        return result or []

    def insert_one(self, table: str, payload: dict[str, Any]) -> dict[str, Any]:
        result = self.request("POST", table, payload=payload, prefer="return=representation")
        if not result:
            raise ImportErrorDetalhado(f"INSERT em {table} não retornou o registro criado.")
        return result[0] if isinstance(result, list) else result


def normalizar_texto(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def converter_data(value: Any, linha: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    texto = normalizar_texto(value)
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date().isoformat()
        except ValueError:
            pass

    raise ImportErrorDetalhado(f"Linha {linha}: data inválida: {value!r}")


def converter_valor(value: Any, linha: int) -> Decimal:
    try:
        valor = Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, AttributeError):
        raise ImportErrorDetalhado(f"Linha {linha}: valor inválido: {value!r}")

    if valor == 0:
        raise ImportErrorDetalhado(f"Linha {linha}: valor não pode ser zero.")

    # Os dados bancários exportados representam despesas como negativas.
    # A tela de Lançamentos recebe o valor da operação como magnitude positiva;
    # o campo 'tipo' determina se ela é RECEITA ou DESPESA.
    return abs(valor).quantize(Decimal("0.01"))


def ler_planilha(caminho: str, nome_aba: str) -> list[RegistroEntrada]:
    workbook = load_workbook(caminho, data_only=True, read_only=True)
    if nome_aba not in workbook.sheetnames:
        raise ImportErrorDetalhado(
            f"A aba {nome_aba!r} não existe. Abas disponíveis: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[nome_aba]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [normalizar_texto(v) for v in next(rows)]
    except StopIteration:
        raise ImportErrorDetalhado("A aba de importação está vazia.")

    if headers != EXPECTED_COLUMNS:
        raise ImportErrorDetalhado(
            "Colunas da aba 'base' não correspondem ao formato esperado.\n"
            f"Esperado: {EXPECTED_COLUMNS}\n"
            f"Encontrado: {headers}"
        )

    registros: list[RegistroEntrada] = []
    for linha, values in enumerate(rows, start=2):
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        data = converter_data(values[0], linha)
        descricao = normalizar_texto(values[1])
        tipo = normalizar_texto(values[3]).upper()
        tipo_pessoa = normalizar_texto(values[4]).upper()
        conta = normalizar_texto(values[5]) or None
        categoria = normalizar_texto(values[6])
        meio_pagamento = normalizar_texto(values[7]).upper()
        cartao = normalizar_texto(values[8]) or None

        try:
            parcelas = int(values[9])
        except (TypeError, ValueError):
            raise ImportErrorDetalhado(f"Linha {linha}: parcelas inválidas: {values[9]!r}")

        if not descricao:
            raise ImportErrorDetalhado(f"Linha {linha}: descrição obrigatória.")
        if tipo not in {"RECEITA", "DESPESA"}:
            raise ImportErrorDetalhado(f"Linha {linha}: tipo inválido: {tipo!r}")
        if tipo_pessoa not in {"PF", "PJ"}:
            raise ImportErrorDetalhado(f"Linha {linha}: tipo_pessoa inválido: {tipo_pessoa!r}")
        if meio_pagamento not in {"CONTA", "CARTAO"}:
            raise ImportErrorDetalhado(f"Linha {linha}: meio_pagamento inválido: {meio_pagamento!r}")
        if parcelas < 1:
            raise ImportErrorDetalhado(f"Linha {linha}: parcelas deve ser >= 1.")
        if meio_pagamento == "CARTAO" and tipo != "DESPESA":
            raise ImportErrorDetalhado(
                f"Linha {linha}: a tela atual força CARTAO a ser DESPESA; receita em cartão não é aceita."
            )
        if meio_pagamento == "CONTA" and not conta:
            raise ImportErrorDetalhado(f"Linha {linha}: conta é obrigatória para meio_pagamento=CONTA.")
        if meio_pagamento == "CARTAO" and not cartao:
            raise ImportErrorDetalhado(f"Linha {linha}: cartao é obrigatório para meio_pagamento=CARTAO.")

        registros.append(
            RegistroEntrada(
                linha=linha,
                data=data,
                descricao=descricao,
                valor=converter_valor(values[2], linha),
                tipo=tipo,
                tipo_pessoa=tipo_pessoa,
                conta=conta,
                categoria=categoria,
                meio_pagamento=meio_pagamento,
                cartao=cartao,
                parcelas=parcelas if meio_pagamento == "CARTAO" else 1,
            )
        )

    return registros


def construir_caminhos_categorias(categorias: list[Categoria]) -> dict[str, list[Categoria]]:
    por_id = {c.id: c for c in categorias}
    caminhos: dict[str, list[Categoria]] = {}

    def caminho(cat: Categoria) -> list[Categoria]:
        atual: list[Categoria] = []
        cursor: Categoria | None = cat
        visitados: set[str] = set()
        while cursor:
            if cursor.id in visitados:
                raise ImportErrorDetalhado(f"Ciclo detectado na hierarquia de categorias em {cursor.id}.")
            visitados.add(cursor.id)
            atual.append(cursor)
            cursor = por_id.get(cursor.parent_id) if cursor.parent_id else None
        atual.reverse()
        return atual

    for cat in categorias:
        partes = caminho(cat)
        nome_caminho = "/".join(c.nome for c in partes)
        caminhos.setdefault(nome_caminho, []).append(cat)
        # Também permite resolver uma categoria por nome simples quando ela for única.
        caminhos.setdefault(cat.nome, []).append(cat)

    return caminhos


def resolver_categoria(nome: str, tipo: str, categorias: list[Categoria], caminhos: dict[str, list[Categoria]]) -> Categoria:
    candidatas = [c for c in caminhos.get(nome, []) if c.tipo == tipo]
    if len(candidatas) == 1:
        return candidatas[0]
    if len(candidatas) > 1:
        raise ImportErrorDetalhado(
            f"Categoria {nome!r} é ambígua para {tipo}. Use o caminho completo (ex.: Pai/Filha)."
        )
    raise ImportErrorDetalhado(f"Categoria {nome!r} ({tipo}) não encontrada no Supabase.")


def encontrar_por_nome(itens: list[dict[str, Any]], campo: str, valor: str, descricao: str) -> dict[str, Any]:
    encontrados = [item for item in itens if normalizar_texto(item.get(campo)) == valor]
    if len(encontrados) == 1:
        return encontrados[0]
    if not encontrados:
        raise ImportErrorDetalhado(f"{descricao} {valor!r} não encontrado.")
    raise ImportErrorDetalhado(f"{descricao} {valor!r} aparece mais de uma vez e é ambíguo.")


def carregar_referencias(db: SupabaseREST) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[Categoria]]:
    contas = db.select("contas_bancarias", "id,apelido")
    cartoes = db.select("cartoes_credito", "id,nome,dia_fechamento,dia_vencimento")
    categorias_raw = db.select("categorias", "id,nome,tipo,parent_id")
    categorias = [
        Categoria(
            id=c["id"],
            nome=c["nome"],
            tipo=c["tipo"].upper(),
            parent_id=c.get("parent_id"),
        )
        for c in categorias_raw
    ]
    return contas, cartoes, categorias


def calcular_mes_ano_fatura(data_compra: str, dia_fechamento: int, parcela_index: int) -> tuple[int, int]:
    compra = date.fromisoformat(data_compra)
    ano = compra.year
    mes = compra.month

    if compra.day > dia_fechamento:
        mes += 1

    mes += parcela_index - 1
    while mes > 12:
        mes -= 12
        ano += 1

    return ano, mes


def validar_referencias(
    registros: list[RegistroEntrada],
    contas: list[dict[str, Any]],
    cartoes: list[dict[str, Any]],
    categorias: list[Categoria],
) -> None:
    caminhos = construir_caminhos_categorias(categorias)
    erros: list[str] = []

    for registro in registros:
        try:
            resolver_categoria(registro.categoria, registro.tipo, categorias, caminhos)
            if registro.meio_pagamento == "CONTA":
                encontrar_por_nome(contas, "apelido", registro.conta or "", "Conta")
            else:
                encontrar_por_nome(cartoes, "nome", registro.cartao or "", "Cartão")
        except ImportErrorDetalhado as exc:
            erros.append(f"Linha {registro.linha}: {exc}")

    if erros:
        raise ImportErrorDetalhado("Foram encontradas referências inválidas:\n- " + "\n- ".join(erros))


def importar(
    db: SupabaseREST,
    registros: list[RegistroEntrada],
    contas: list[dict[str, Any]],
    cartoes: list[dict[str, Any]],
    categorias: list[Categoria],
    criado_por_nome: str,
) -> None:
    caminhos = construir_caminhos_categorias(categorias)

    for indice, registro in enumerate(registros, start=1):
        categoria = resolver_categoria(registro.categoria, registro.tipo, categorias, caminhos)

        if registro.meio_pagamento == "CONTA":
            conta = encontrar_por_nome(contas, "apelido", registro.conta or "", "Conta")
            conta_id = conta["id"]
            cartao = None
        else:
            cartao = encontrar_por_nome(cartoes, "nome", registro.cartao or "", "Cartão")
            # Reproduz exatamente o comportamento atual de page.tsx:
            # conta_id do lançamento recebe o ID do cartão selecionado.
            conta_id = cartao["id"]

        lancamento = db.insert_one(
            "lancamentos",
            {
                "descricao": registro.descricao,
                "valor": float(registro.valor),
                "data": registro.data,
                "tipo": registro.tipo,
                "categoria_id": categoria.id,
                "conta_id": conta_id,
                "meio_pagamento": registro.meio_pagamento,
                "parcelas": registro.parcelas,
                "tipo_pessoa": registro.tipo_pessoa,
                "criado_por_nome": criado_por_nome,
            },
        )

        lancamento_id = lancamento["id"]

        if registro.meio_pagamento == "CONTA":
            db.insert_one(
                "transacoes",
                {
                    "lancamento_id": lancamento_id,
                    "descricao": registro.descricao,
                    "valor": float(registro.valor),
                    "tipo": registro.tipo.lower(),
                    "status": "pago",
                    "data_competencia": registro.data,
                    "categoria_id": categoria.id,
                    "conta_bancaria_id": conta_id,
                    "numero_parcela": 1,
                    "total_parcelas": 1,
                    "criado_por_nome": criado_por_nome,
                },
            )
        else:
            valor_parcela = (registro.valor / registro.parcelas).quantize(Decimal("0.01"))

            for parcela in range(1, registro.parcelas + 1):
                ano, mes = calcular_mes_ano_fatura(
                    registro.data,
                    int(cartao["dia_fechamento"]),
                    parcela,
                )

                faturas = db.select(
                    "faturas",
                    "id",
                    {
                        "cartao_credito_id": f"eq.{cartao['id']}",
                        "ano": f"eq.{ano}",
                        "mes": f"eq.{mes}",
                    },
                )

                if faturas:
                    fatura_id = faturas[0]["id"]
                else:
                    fatura = db.insert_one(
                        "faturas",
                        {
                            "cartao_credito_id": cartao["id"],
                            "ano": ano,
                            "mes": mes,
                            "status": "ABERTA",
                        },
                    )
                    fatura_id = fatura["id"]

                data_competencia = f"{ano:04d}-{mes:02d}-{int(cartao['dia_vencimento']):02d}"
                descricao = (
                    f"{registro.descricao} ({parcela}/{registro.parcelas})"
                    if registro.parcelas > 1
                    else registro.descricao
                )

                db.insert_one(
                    "transacoes",
                    {
                        "lancamento_id": lancamento_id,
                        "fatura_id": fatura_id,
                        "cartao_credito_id": cartao["id"],
                        "descricao": descricao,
                        "valor": float(valor_parcela),
                        "tipo": "despesa",
                        "status": "pendente",
                        "data_competencia": data_competencia,
                        "categoria_id": categoria.id,
                        "numero_parcela": parcela,
                        "total_parcelas": registro.parcelas,
                        "criado_por_nome": criado_por_nome,
                    },
                )

        print(f"[{indice}/{len(registros)}] importado: {registro.data} - {registro.descricao}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa a aba base para o fluxo de Lançamentos.")
    parser.add_argument("arquivo", help="Caminho da planilha XLSX")
    parser.add_argument("--aba", default="base", help="Aba a importar (padrão: base)")
    parser.add_argument(
        "--executar",
        action="store_true",
        help="Efetivamente grava no Supabase. Sem esta opção, apenas valida.",
    )
    args = parser.parse_args()

    try:
        registros = ler_planilha(args.arquivo, args.aba)
        print(f"Arquivo: {args.arquivo}")
        print(f"Aba: {args.aba}")
        print(f"Lançamentos encontrados: {len(registros)}")

        url = os.getenv("SUPABASE_URL")
        key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        if not url or not key:
            raise ImportErrorDetalhado(
                "Defina SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY antes de executar o importador."
            )

        db = SupabaseREST(url, key)
        contas, cartoes, categorias = carregar_referencias(db)
        validar_referencias(registros, contas, cartoes, categorias)

        print("Validação das referências: OK")
        print(f"Contas disponíveis: {len(contas)}")
        print(f"Cartões disponíveis: {len(cartoes)}")
        print(f"Categorias disponíveis: {len(categorias)}")
        print("Valores de despesas serão convertidos de negativos para positivos; o campo tipo continuará determinando DESPESA/RECEITA.")

        if not args.executar:
            print("\nDRY-RUN: nenhum dado foi gravado. Use --executar para realizar a carga.")
            return 0

        criado_por_nome = os.getenv("IMPORT_CRIADO_POR_NOME", "Importação").strip() or "Importação"
        print(f"Iniciando carga com criado_por_nome={criado_por_nome!r}...")
        importar(db, registros, contas, cartoes, categorias, criado_por_nome)
        print("\nCarga concluída.")
        return 0

    except ImportErrorDetalhado as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
